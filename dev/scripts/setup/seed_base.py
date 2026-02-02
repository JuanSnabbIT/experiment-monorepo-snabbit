#!/usr/bin/env python
"""
Script orquestador que ejecuta scripts de seed en orden correcto.

Que hace:
- Ejecuta setup_superuser.py (si no existe superusuario) y luego pobla datos base completos
- Verifica exito de cada paso antes de continuar
- Proporciona resumen al finalizar
- Crea SOLO datos base (NO crea flujos transaccionales)

IMPORTANTE: Este script pobla datos base para permitir pruebas manuales:
- Empresas (con recargo entre 22-28% y PPM entre 3-7%)
- Sucursales (con datos completos de región, provincia, comuna)
- Usuarios internos y de clientes
- Catálogos (Items, Proveedores, Categorías, Servicios, etc.)
- Stock inicial en bodegas

NO crea automáticamente:
- Cotizaciones
- Órdenes de Trabajo
- Contratos
- Órdenes de Compra
- Compras
- Guías de Salida

Estos registros deben crearse manualmente para realizar pruebas reales del sistema.

Cuando usar:
- Despues de limpiar la base de datos para poblar sistema desde cero
- Primera inicializacion del sistema
- Testing con datos base solidos

Prerequisitos:
- Base de datos limpia (por ejemplo, borrar db.sqlite3)
- Migraciones aplicadas (manage.py migrate)

Uso:
    cd backend
    backend\\ENV\\Scripts\\python.exe ..\\dev\\scripts\\setup\\seed_base.py
"""
from __future__ import annotations

import os
import random
import re
import subprocess
import sys
import unicodedata
from datetime import date, timedelta
from pathlib import Path
from typing import Dict, List, Optional

# Rutas
REPO_ROOT = Path(__file__).resolve().parents[3]
BACKEND_PATH = REPO_ROOT / "backend"
SCRIPTS_DIR = Path(__file__).resolve().parent
PYTHON_EXE = BACKEND_PATH / "ENV" / "Scripts" / "python.exe"

# Scripts a ejecutar en orden de dependencias
SEED_SCRIPTS = [
    {
        "name": "setup_superuser.py",
        "description": "Crear superusuario, empresa base (11111111-1) y grupos",
        "required": True,
        "condition": "no_superuser",  # Solo ejecutar si no hay superusuario
    },
    {
        "name": "seed_base.py",
        "description": "Poblar datos base completos",
        "required": True,
        "condition": None,  # Siempre ejecutar
    },
]

DJANGO_READY = False


def print_header(text: str) -> None:
    """Imprime encabezado con formato."""
    print("\n" + "=" * 80)
    print(text.center(80))
    print("=" * 80)


def print_step(
    step_num: int, total_steps: int, script_name: str, description: str
) -> None:
    """Imprime informacion del paso actual."""
    print(f"\n{'=' * 80}")
    print(f"PASO {step_num}/{total_steps}: {script_name}")
    print(f"Descripcion: {description}")
    print(f"{'=' * 80}")


def python_executable() -> str:
    """Resuelve el Python del entorno local o usa sys.executable."""
    if PYTHON_EXE.exists():
        return str(PYTHON_EXE)
    return sys.executable


def bootstrap_django() -> None:
    """Inicializa el entorno Django."""
    global DJANGO_READY
    if DJANGO_READY:
        return

    backend_path = str(BACKEND_PATH)
    if backend_path not in sys.path:
        sys.path.insert(0, backend_path)
    try:
        os.chdir(backend_path)
    except Exception:
        pass

    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "sw_erp.settings")
    import django

    django.setup()
    DJANGO_READY = True


def ensure_migrations() -> None:
    """Aplica migraciones (y crea si faltan) para las apps principales."""
    from django.core.management import call_command

    apps_to_migrate = [
        "core",
        "cuentas",
        "empresas",
        "items",
        "bodegas",
        "recursos",
        "cotizaciones",
        "contratos",
        "visitas",
        "calendario",
        "activos",
        "rendiciones",
        # "retroalimentacion",  # DESACTIVADA - Migrada a ordentrabajov2
        "vacaciones",
        "ordentrabajov2",
    ]
    print("=> Ejecutando makemigrations (si aplica) y migrate...")
    call_command("makemigrations", *apps_to_migrate, "--noinput")
    call_command("migrate", "--noinput")


def check_condition(condition: Optional[str]) -> bool:
    """Verifica si se debe ejecutar el script basado en condicion."""
    if condition is None:
        return True
    if condition == "no_superuser":
        bootstrap_django()
        from django.contrib.auth import get_user_model

        User = get_user_model()
        return not User.objects.filter(is_superuser=True).exists()
    return True


def run_script(script_path: Path) -> bool:
    """Ejecuta un script y retorna True si tuvo exito."""
    args = [python_executable(), str(script_path)]
    if script_path.name == "seed_base.py":
        args.append("--internal")
    result = subprocess.run(
        args,
        cwd=str(BACKEND_PATH),
        capture_output=False,
    )
    return result.returncode == 0


def normalize_header(value: Optional[str]) -> str:
    """Normaliza headers de Excel para matching flexible."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = "".join(
        ch
        for ch in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(ch)
    )
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def slugify(value: str) -> str:
    return normalize_header(value).replace(" ", "")


def load_excel_rows(file_path: Path) -> List[Dict[str, object]]:
    """Carga filas de un Excel (pandas/openpyxl) y normaliza headers."""
    if not file_path.exists():
        return []

    try:
        import pandas as pd  # type: ignore

        df = pd.read_excel(file_path)
        df.columns = [normalize_header(col) for col in df.columns]
        df = df.where(pd.notnull(df), None)
        return df.to_dict(orient="records")
    except ImportError:
        pass
    except Exception as exc:
        print(f"Advertencia: fallo leyendo {file_path.name} con pandas: {exc}")

    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [normalize_header(h) for h in rows[0]]
        data = []
        for row in rows[1:]:
            if not any(row):
                continue
            row_dict: Dict[str, object] = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    row_dict[headers[idx]] = value
            data.append(row_dict)
        return data
    except ImportError:
        print("Advertencia: openpyxl no esta instalado, se omite lectura Excel.")
    except Exception as exc:
        print(f"Advertencia: fallo leyendo {file_path.name} con openpyxl: {exc}")
    return []


def pick_value(row: Dict[str, object], candidates: List[str]) -> Optional[str]:
    """Busca el primer valor de una lista de columnas candidatas."""
    for candidate in candidates:
        for key, value in row.items():
            if not key:
                continue
            if key == candidate or candidate in key:
                if value is None:
                    continue
                return str(value).strip()
    return None


def clean_rut(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def split_full_name(full_name: str) -> Dict[str, str]:
    parts = [p for p in full_name.strip().split(" ") if p]
    if not parts:
        return {"first_name": "Usuario", "last_name": "Cliente"}
    if len(parts) == 1:
        return {"first_name": parts[0], "last_name": "Cliente"}
    if len(parts) == 2:
        return {"first_name": parts[0], "last_name": parts[1]}
    return {
        "first_name": parts[0],
        "second_name": " ".join(parts[1:-1]),
        "last_name": parts[-1],
    }


def build_rut_generator(start: int = 20000000):
    counter = start

    def next_rut() -> str:
        nonlocal counter
        from django.contrib.auth import get_user_model

        User = get_user_model()
        while True:
            rut = f"{counter}-1"
            counter += 1
            if not User.objects.filter(rut=rut).exists():
                return rut

    return next_rut


def ensure_user(
    email: str,
    password: str,
    first_name: str,
    last_name: str,
    rut: str,
    *,
    is_active: bool = True,
    is_staff: bool = False,
    is_superuser: bool = False,
    second_name: Optional[str] = None,
    second_last_name: Optional[str] = None,
) -> object:
    from django.contrib.auth import get_user_model

    User = get_user_model()
    user = User.objects.filter(email=email).first()
    if not user:
        user = User.objects.create_user(
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            rut=rut,
            second_name=second_name,
            second_last_name=second_last_name,
        )
    updated = False
    if user.is_active != is_active:
        user.is_active = is_active
        updated = True
    if is_staff and not user.is_staff:
        user.is_staff = True
        updated = True
    if is_superuser and not user.is_superuser:
        user.is_superuser = True
        updated = True
    if first_name and user.first_name != first_name:
        user.first_name = first_name
        updated = True
    if last_name and user.last_name != last_name:
        user.last_name = last_name
        updated = True
    if second_name is not None and user.second_name != second_name:
        user.second_name = second_name
        updated = True
    if second_last_name is not None and user.second_last_name != second_last_name:
        user.second_last_name = second_last_name
        updated = True
    if rut and user.rut != rut:
        user.rut = rut
        updated = True
    if updated:
        user.save()
    return user


def ensure_personalizacion(user: object, sucursal: object) -> None:
    from core.models import PersonalizacionUsuario

    personalizacion, _ = PersonalizacionUsuario.objects.get_or_create(
        usuario=user,
        defaults={"tema": "3", "font_size": 14, "sucursal_principal": sucursal},
    )
    if personalizacion.sucursal_principal_id != sucursal.id:
        personalizacion.sucursal_principal = sucursal
        personalizacion.save(update_fields=["sucursal_principal"])


def ensure_usuario_empresa(
    user: object,
    sucursal: object,
    grupos: Optional[List[object]] = None,
    *,
    cargo: Optional[str] = None,
    fecha_contrato: Optional[date] = None,
) -> object:
    from empresas.models import UsuarioEmpresa

    defaults = {"sucursal": sucursal, "estado": "1"}
    if cargo:
        defaults["cargo"] = cargo
    if fecha_contrato:
        defaults["fecha_contrato"] = fecha_contrato
        defaults["fecha_ingreso"] = fecha_contrato

    usuario_empresa, _created = UsuarioEmpresa.objects.get_or_create(
        usuario=user,
        defaults=defaults,
    )
    updated = False
    if usuario_empresa.sucursal_id != sucursal.id:
        usuario_empresa.sucursal = sucursal
        updated = True
    if usuario_empresa.estado != "1":
        usuario_empresa.estado = "1"
        updated = True
    if cargo and usuario_empresa.cargo != cargo:
        usuario_empresa.cargo = cargo
        updated = True
    if fecha_contrato and usuario_empresa.fecha_contrato != fecha_contrato:
        usuario_empresa.fecha_contrato = fecha_contrato
        usuario_empresa.fecha_ingreso = fecha_contrato
        updated = True
    if updated:
        usuario_empresa.save()
    if grupos is not None:
        usuario_empresa.grupos.set(grupos)
    return usuario_empresa


def ensure_groups() -> Dict[str, object]:
    from django.contrib.auth.models import Group

    grupos_def = [
        ("staff", "Personal administrativo general"),
        ("superadmin", "Administrador con permisos maximos"),
        ("multi-empresas", "Acceso a multiples empresas"),
        ("tecnico", "Tecnico de campo para OT y visitas"),
        ("bodeguero", "Gestores de inventario y bodegas"),
        ("representante_legal", "Representante legal de empresa"),
    ]
    grupos: Dict[str, object] = {}
    for nombre, _descripcion in grupos_def:
        grupo, _ = Group.objects.get_or_create(name=nombre)
        grupos[nombre] = grupo
    return grupos


def ensure_empresa_base() -> tuple[object, object]:
    from empresas.models import Empresa, SucursalEmpresa
    from decimal import Decimal

    # Generar recargo entre 22% y 28%
    recargo = random.randint(22, 28)
    # Generar ppm entre 3% y 7% (como decimal)
    ppm = Decimal(str(random.uniform(3.0, 7.0)))

    empresa, _ = Empresa.objects.get_or_create(
        rut_empresa="11111111-1",
        defaults={
            "nombre": "Snabbit",
            "direccion_principal": "Av. Providencia 1234, Santiago",
            "telefono": "+56912345678",
            "email": "contacto@snabbit.cl",
            "recargo": recargo,
            "ppm": ppm,
        },
    )
    updated = False
    if not empresa.nombre:
        empresa.nombre = "Snabbit"
        updated = True
    if not empresa.direccion_principal:
        empresa.direccion_principal = "Av. Providencia 1234, Santiago"
        updated = True
    if empresa.recargo == 0:
        empresa.recargo = recargo
        updated = True
    if empresa.ppm == 1:
        empresa.ppm = ppm
        updated = True
    if updated:
        empresa.save()

    sucursal, _ = SucursalEmpresa.objects.get_or_create(
        empresa=empresa,
        nombre="Casa Matriz",
        defaults={
            "direccion": empresa.direccion_principal,
            "telefono": empresa.telefono,
            "email": empresa.email,
            "region": 13,  # Región Metropolitana
            "provincia": 131,  # Santiago
            "comuna": 13101,  # Santiago Centro
        },
    )
    # Actualizar sucursal si le faltan datos de ubicación
    updated_suc = False
    if sucursal.region == 0:
        sucursal.region = 13
        updated_suc = True
    if sucursal.provincia == 0:
        sucursal.provincia = 131
        updated_suc = True
    if sucursal.comuna == 0:
        sucursal.comuna = 13101
        updated_suc = True
    if updated_suc:
        sucursal.save()

    return empresa, sucursal


def ensure_internal_users(
    empresa: object,
    sucursal: object,
    grupos: Dict[str, object],
) -> Dict[str, object]:
    from django.contrib.auth import get_user_model

    User = get_user_model()
    has_superuser = User.objects.filter(is_superuser=True).exists()
    admin_superuser = not has_superuser

    internal_specs = [
        {
            "key": "admin",
            "email": "admin@snabbit.cl",
            "rut": "11111111-9",
            "first_name": "Admin",
            "last_name": "Snabbit",
            "password": "test1234",
            "is_staff": True,
            "is_superuser": admin_superuser,
            "grupos": ["staff", "superadmin", "multi-empresas"],
            "cargo": "Administrador",
        },
        {
            "key": "ventas",
            "email": "ventas@snabbit.cl",
            "rut": "11111111-6",
            "first_name": "Paula",
            "last_name": "Ventas",
            "password": "test1234",
            "is_staff": True,
            "is_superuser": False,
            "grupos": ["staff"],
            "cargo": "Ejecutiva de Ventas",
        },
        {
            "key": "contabilidad",
            "email": "contabilidad@snabbit.cl",
            "rut": "11111111-5",
            "first_name": "Diego",
            "last_name": "Contabilidad",
            "password": "test1234",
            "is_staff": True,
            "is_superuser": False,
            "grupos": ["staff"],
            "cargo": "Analista Contable",
        },
        {
            "key": "supervisor",
            "email": "supervisor@snabbit.cl",
            "rut": "11111111-4",
            "first_name": "Camila",
            "last_name": "Supervisor",
            "password": "test1234",
            "is_staff": True,
            "is_superuser": False,
            "grupos": ["staff"],
            "cargo": "Supervisora de Operaciones",
        },
        {
            "key": "tecnico",
            "email": "tecnico@snabbit.cl",
            "rut": "11111111-8",
            "first_name": "Tecnico",
            "last_name": "Snabbit",
            "password": "test1234",
            "is_staff": False,
            "is_superuser": False,
            "grupos": ["tecnico"],
            "cargo": "Tecnico",
        },
        {
            "key": "bodeguero",
            "email": "bodeguero@snabbit.cl",
            "rut": "11111111-7",
            "first_name": "Bodeguero",
            "last_name": "Snabbit",
            "password": "test1234",
            "is_staff": False,
            "is_superuser": False,
            "grupos": ["bodeguero"],
            "cargo": "Bodeguero",
        },
    ]

    results: Dict[str, object] = {}
    for spec in internal_specs:
        user = ensure_user(
            email=spec["email"],
            password=spec["password"],
            first_name=spec["first_name"],
            last_name=spec["last_name"],
            rut=spec["rut"],
            is_active=True,
            is_staff=spec["is_staff"],
            is_superuser=spec["is_superuser"],
        )
        grupos_usuario = [grupos[name] for name in spec["grupos"] if name in grupos]
        usuario_empresa = ensure_usuario_empresa(
            user,
            sucursal,
            grupos_usuario,
            cargo=spec["cargo"],
            fecha_contrato=date.today() - timedelta(days=400),
        )
        ensure_personalizacion(user, sucursal)
        results[spec["key"]] = usuario_empresa
    return results


def extract_users_from_excel(file_path: Path, rut_generator) -> List[Dict[str, str]]:
    rows = load_excel_rows(file_path)
    if not rows:
        return []

    users: List[Dict[str, str]] = []
    seen_emails = set()
    for row in rows:
        email = pick_value(row, ["email", "correo", "correo electronico", "mail"])
        if not email:
            continue
        email = email.strip().lower()
        if email in seen_emails:
            continue

        rut = clean_rut(pick_value(row, ["rut", "run"]))
        nombre = pick_value(row, ["nombre", "nombres"])
        apellido = pick_value(row, ["apellido", "apellidos"])
        nombre_completo = pick_value(
            row,
            ["nombre completo", "nombre y apellido", "nombre apellido"],
        )
        cargo = pick_value(row, ["cargo", "puesto"])
        second_name = pick_value(row, ["segundo nombre"])
        second_last_name = pick_value(row, ["segundo apellido"])

        if not (nombre and apellido) and nombre_completo:
            parts = split_full_name(nombre_completo)
            nombre = parts.get("first_name")
            apellido = parts.get("last_name")
            if not second_name:
                second_name = parts.get("second_name")

        if not nombre:
            nombre = "Usuario"
        if not apellido:
            apellido = "Cliente"

        if not rut:
            rut = rut_generator()

        users.append(
            {
                "email": email,
                "rut": rut,
                "first_name": nombre,
                "last_name": apellido,
                "second_name": second_name,
                "second_last_name": second_last_name,
                "cargo": cargo,
            }
        )
        seen_emails.add(email)
    return users


def ensure_client_companies(
    prestador: object,
    grupos: Dict[str, object],
) -> List[Dict[str, object]]:
    from empresas.models import Empresa, RelacionEmpresa, SucursalEmpresa

    clients_config = [
        {
            "nombre": "AYG ASOCIADOS SpA",
            "rut": "76123456-7",
            "direccion": "Av. Providencia 1234, Oficina 402",
            "telefono": "+56223456789",
            "email": "contacto@aygasociados.cl",
            "region": 13,
            "provincia": 131,
            "comuna": 13101,
            "archivo": "usuarios_aygasociados.xlsx",
        },
        {
            "nombre": "CAMACOES Ltda",
            "rut": "76345678-5",
            "direccion": "Av. Libertad 450, Piso 3",
            "telefono": "+56322567890",
            "email": "contacto@camacoes.cl",
            "region": 5,
            "provincia": 51,
            "comuna": 5101,
            "archivo": "usuarios_camacoes.xlsx",
        },
        {
            "nombre": "MOLINA RIOS Ingeniería",
            "rut": "76456789-0",
            "direccion": "Av. San Martin 890",
            "telefono": "+56422345678",
            "email": "contacto@molinarios.cl",
            "region": 8,
            "provincia": 81,
            "comuna": 8101,
            "archivo": "usuarios_molinarios.xlsx",
        },
        {
            "nombre": "PRODALMEN S.A.",
            "rut": "76567890-1",
            "direccion": "Calle 10 Norte 120",
            "telefono": "+56322560123",
            "email": "contacto@prodalmen.cl",
            "region": 7,
            "provincia": 71,
            "comuna": 7101,
            "archivo": "usuarios_prodalmen.xlsx",
        },
        {
            "nombre": "Segurimax Chile SpA",
            "rut": "76678901-2",
            "direccion": "Av. Apoquindo 4890, Torre B",
            "telefono": "+56221234567",
            "email": "contacto@segurimax.cl",
            "region": 13,
            "provincia": 131,
            "comuna": 13114,
            "archivo": "usuarios_segurimax.xlsx",
        },
        {
            "nombre": "Servicios Nova Ltda",
            "rut": "76789012-3",
            "direccion": "Camino Industrial 455",
            "telefono": "+56512223344",
            "email": "contacto@serviciosnova.cl",
            "region": 2,
            "provincia": 21,
            "comuna": 2101,
            "archivo": "usuarios_serviciosnova.xlsx",
        },
        {
            "nombre": "Grupo Austral SpA",
            "rut": "76890123-4",
            "direccion": "Pedro Montt 1200",
            "telefono": "+56652223344",
            "email": "contacto@grupoaustral.cl",
            "region": 10,
            "provincia": 101,
            "comuna": 10101,
            "archivo": "usuarios_grupoaustral.xlsx",
        },
        {
            "nombre": "Centro Medico San Lucas",
            "rut": "76901234-5",
            "direccion": "Av. Alemania 850",
            "telefono": "+56452345678",
            "email": "contacto@cmsanlucas.cl",
            "region": 9,
            "provincia": 91,
            "comuna": 9101,
            "archivo": "usuarios_cmsanlucas.xlsx",
        },
        {
            "nombre": "Logistica Rayo Sur",
            "rut": "77012345-6",
            "direccion": "Parque Industrial 200",
            "telefono": "+56342223344",
            "email": "contacto@rayosur.cl",
            "region": 6,
            "provincia": 61,
            "comuna": 6101,
            "archivo": "usuarios_rayosur.xlsx",
        },
        {
            "nombre": "Fundacion Valle Seguro",
            "rut": "77123456-7",
            "direccion": "Av. O'Higgins 322",
            "telefono": "+56712345678",
            "email": "contacto@valleseguro.cl",
            "region": 3,
            "provincia": 31,
            "comuna": 3101,
            "archivo": "usuarios_valleseguro.xlsx",
        },
        {
            "nombre": "Comercial El Faro EIRL",
            "rut": "77234567-8",
            "direccion": "Av. La Serena 120",
            "telefono": "+56512229900",
            "email": "contacto@comercialelfaro.cl",
            "region": 4,
            "provincia": 41,
            "comuna": 4101,
            "archivo": "usuarios_comercialelfaro.xlsx",
        },
        {
            "nombre": "Juan Perez Servicios",
            "rut": "77345678-9",
            "direccion": "Los Jardines 45",
            "telefono": "+56229876543",
            "email": "contacto@juanperezservicios.cl",
            "region": 13,
            "provincia": 132,
            "comuna": 13201,
            "archivo": "usuarios_juanperez.xlsx",
        },
    ]

    from django.contrib.auth import get_user_model

    User = get_user_model()
    rut_generator = build_rut_generator()
    existing_emails = set(User.objects.values_list("email", flat=True))

    results: List[Dict[str, object]] = []
    for client in clients_config:
        # Generar recargo entre 22% y 28% para cada empresa cliente
        recargo = random.randint(22, 28)
        # Generar ppm entre 3% y 7% (como decimal)
        from decimal import Decimal
        ppm = Decimal(str(random.uniform(3.0, 7.0)))

        empresa, _ = Empresa.objects.get_or_create(
            rut_empresa=client["rut"],
            defaults={
                "nombre": client["nombre"],
                "direccion_principal": client.get("direccion") or "Direccion Cliente",
                "telefono": client.get("telefono") or "+56900000000",
                "email": client.get("email")
                or f"contacto@{slugify(client['nombre'])}.cl",
                "recargo": recargo,
                "ppm": ppm,
            },
        )

        # Actualizar recargo y ppm si están en valores por defecto
        updated = False
        if empresa.recargo == 0:
            empresa.recargo = recargo
            updated = True
        if empresa.ppm == 1:
            empresa.ppm = ppm
            updated = True
        if updated:
            empresa.save()

        sucursal, _ = SucursalEmpresa.objects.get_or_create(
            empresa=empresa,
            nombre="Casa Matriz",
            defaults={
                "direccion": empresa.direccion_principal,
                "telefono": empresa.telefono,
                "email": empresa.email,
                "region": client.get("region", 0),
                "provincia": client.get("provincia", 0),
                "comuna": client.get("comuna", 0),
            },
        )

        RelacionEmpresa.objects.get_or_create(
            prestador_servicios=prestador,
            cliente=empresa,
        )

        excel_path = BACKEND_PATH / client["archivo"]
        usuarios_data = extract_users_from_excel(excel_path, rut_generator)
        if not usuarios_data:
            slug = slugify(client["nombre"])
            usuarios_data = [
                {
                    "email": f"admin@{slug}.cl",
                    "rut": rut_generator(),
                    "first_name": "Admin",
                    "last_name": client["nombre"].split(" ")[0].title(),
                    "second_name": None,
                    "second_last_name": None,
                    "cargo": "Administrador",
                },
                {
                    "email": f"compras@{slug}.cl",
                    "rut": rut_generator(),
                    "first_name": "Compras",
                    "last_name": client["nombre"].split(" ")[0].title(),
                    "second_name": None,
                    "second_last_name": None,
                    "cargo": "Compras",
                },
                {
                    "email": f"operaciones@{slug}.cl",
                    "rut": rut_generator(),
                    "first_name": "Operaciones",
                    "last_name": client["nombre"].split(" ")[0].title(),
                    "second_name": None,
                    "second_last_name": None,
                    "cargo": "Operaciones",
                },
            ]

        usuarios_empresa: List[object] = []
        for idx, data in enumerate(usuarios_data):
            email = data["email"].lower()
            grupos_usuario = []
            if idx == 0 and "representante_legal" in grupos:
                grupos_usuario = [grupos["representante_legal"]]

            if email in existing_emails:
                user = User.objects.filter(email=email).first()
                if not user:
                    continue
                from empresas.models import UsuarioEmpresa

                existente = UsuarioEmpresa.objects.filter(usuario=user).first()
                if existente and existente.sucursal.empresa_id != empresa.id:
                    continue
                usuario_empresa = ensure_usuario_empresa(
                    user,
                    sucursal,
                    grupos_usuario,
                    cargo=data.get("cargo"),
                    fecha_contrato=date.today() - timedelta(days=400),
                )
                ensure_personalizacion(user, sucursal)
                usuarios_empresa.append(usuario_empresa)
                continue

            rut = data["rut"]
            if User.objects.filter(rut=rut).exists():
                rut = rut_generator()

            user = ensure_user(
                email=email,
                password="test1234",
                first_name=data["first_name"],
                last_name=data["last_name"],
                rut=rut,
                is_active=True,
                is_staff=False,
                second_name=data.get("second_name"),
                second_last_name=data.get("second_last_name"),
            )

            usuario_empresa = ensure_usuario_empresa(
                user,
                sucursal,
                grupos_usuario,
                cargo=data.get("cargo"),
                fecha_contrato=date.today() - timedelta(days=400),
            )
            ensure_personalizacion(user, sucursal)
            usuarios_empresa.append(usuario_empresa)
            existing_emails.add(email)

        results.append(
            {
                "empresa": empresa,
                "sucursal": sucursal,
                "usuarios": usuarios_empresa,
            }
        )
    return results


def ensure_bodegas(empresa: object, sucursal: object) -> Dict[str, object]:
    from bodegas.models import Bodega

    bodegas = {}
    for nombre in ["Bodega Principal", "Bodega Secundaria"]:
        bodega, _ = Bodega.objects.get_or_create(
            nombre=nombre,
            sucursal=sucursal,
        )
        bodegas[nombre] = bodega
    return bodegas


def ensure_catalogs(empresa: object):
    from items.models import Categoria, Fabricante, ProveedorEmpresa

    categorias = {}
    for nombre in [
        "Camaras IP",
        "Camaras Analogas",
        "Grabadores y NVR",
        "Control de Acceso",
        "Alarmas e Intrusion",
        "Redes y Conectividad",
        "Energia y UPS",
        "Accesorios y Montaje",
        "Cableado y Conectores",
    ]:
        categoria, _ = Categoria.objects.get_or_create(nombre=nombre)
        categorias[nombre] = categoria

    fabricantes = {}
    for nombre in [
        "Hikvision",
        "Dahua",
        "Ubiquiti",
        "TP-Link",
        "Intelbras",
        "APC",
        "Cisco",
        "Genérico",
    ]:
        fabricante, _ = Fabricante.objects.get_or_create(nombre=nombre)
        fabricantes[nombre] = fabricante

    proveedores = []
    proveedores_data = [
        {
            "rut": "76555666-7",
            "nombre": "Importadora TechPro",
            "tipo_moneda": "1",  # USD
            "ejecutivo_asignado": "Juan Perez",
            "email_ejecutivo": "j.perez@techpro.cl",
        },
        {
            "rut": "77888999-0",
            "nombre": "Distribuidora ElectroSur",
            "tipo_moneda": "2",  # CLP
            "ejecutivo_asignado": "Maria Gonzalez",
            "email_ejecutivo": "m.gonzalez@electrosur.cl",
        },
        {
            "rut": "78111222-3",
            "nombre": "Global Hardware Inc",
            "tipo_moneda": "3",  # UF
            "ejecutivo_asignado": "Robert Smith",
            "email_ejecutivo": "r.smith@globalhardware.com",
        },
        {
            "rut": "78222333-4",
            "nombre": "Seguridad Integral Norte",
            "tipo_moneda": "2",
            "ejecutivo_asignado": "Patricia Silva",
            "email_ejecutivo": "p.silva@sintnorte.cl",
        },
        {
            "rut": "78333444-5",
            "nombre": "CCTV Solutions",
            "tipo_moneda": "1",
            "ejecutivo_asignado": "Carlos Diaz",
            "email_ejecutivo": "c.diaz@cctvsolutions.com",
        },
        {
            "rut": "78444555-6",
            "nombre": "Redes y Datos Chile",
            "tipo_moneda": "2",
            "ejecutivo_asignado": "Sofia Romero",
            "email_ejecutivo": "s.romero@redesdatos.cl",
        },
        {
            "rut": "78555666-7",
            "nombre": "Control Access SpA",
            "tipo_moneda": "2",
            "ejecutivo_asignado": "Javier Soto",
            "email_ejecutivo": "j.soto@controlaccess.cl",
        },
        {
            "rut": "78666777-8",
            "nombre": "Energia Segura Ltda",
            "tipo_moneda": "3",
            "ejecutivo_asignado": "Paula Rivas",
            "email_ejecutivo": "p.rivas@energiasegura.cl",
        },
        {
            "rut": "78777888-9",
            "nombre": "Conectividad Andes",
            "tipo_moneda": "2",
            "ejecutivo_asignado": "Felipe Morales",
            "email_ejecutivo": "f.morales@conectividadandes.cl",
        },
        {
            "rut": "78888999-1",
            "nombre": "Accesorios Omega",
            "tipo_moneda": "2",
            "ejecutivo_asignado": "Andrea Lara",
            "email_ejecutivo": "a.lara@accesoriosomega.cl",
        },
    ]

    for data in proveedores_data:
        proveedor, _ = ProveedorEmpresa.objects.get_or_create(
            rut=data["rut"],
            empresa=empresa,
            defaults={
                "nombre": data["nombre"],
                "direccion": "Direccion Proveedor 123",
                "telefono": "+56911111111",
                "email_ejecutivo": data["email_ejecutivo"],
                "ejecutivo_asignado": data["ejecutivo_asignado"],
                "tipo_moneda": data["tipo_moneda"],
            },
        )
        proveedores.append(proveedor)
    return categorias, fabricantes, proveedores


def ensure_items_and_stock(
    empresa: object,
    bodegas: Dict[str, object],
    categorias: Dict[str, object],
    fabricantes: Dict[str, object],
    proveedores: List[object],
) -> List[object]:
    from bodegas.models import StockItemEnBodega
    from items.models import ItemEmpresa

    bodega_principal = bodegas["Bodega Principal"]
    bodega_secundaria = bodegas["Bodega Secundaria"]

    items_data = [
        {
            "nombre": "Camara IP Domo 2MP",
            "categoria": categorias["Camaras IP"],
            "fabricante": fabricantes["Hikvision"],
            "cantidad": 40,
            "bodega": bodega_principal,
            "proveedores": [proveedores[0], proveedores[3]],
        },
        {
            "nombre": "Camara IP Bullet 4MP",
            "categoria": categorias["Camaras IP"],
            "fabricante": fabricantes["Hikvision"],
            "cantidad": 35,
            "bodega": bodega_principal,
            "proveedores": [proveedores[0], proveedores[4]],
        },
        {
            "nombre": "Camara IP Domo 4MP",
            "categoria": categorias["Camaras IP"],
            "fabricante": fabricantes["Dahua"],
            "cantidad": 30,
            "bodega": bodega_principal,
            "proveedores": [proveedores[1], proveedores[4]],
        },
        {
            "nombre": "Camara IP PTZ 20x",
            "categoria": categorias["Camaras IP"],
            "fabricante": fabricantes["Dahua"],
            "cantidad": 12,
            "bodega": bodega_principal,
            "proveedores": [proveedores[1], proveedores[4]],
        },
        {
            "nombre": "Camara Analogica 2MP",
            "categoria": categorias["Camaras Analogas"],
            "fabricante": fabricantes["Hikvision"],
            "cantidad": 25,
            "bodega": bodega_principal,
            "proveedores": [proveedores[0]],
        },
        {
            "nombre": "Camara Analogica 5MP",
            "categoria": categorias["Camaras Analogas"],
            "fabricante": fabricantes["Dahua"],
            "cantidad": 18,
            "bodega": bodega_principal,
            "proveedores": [proveedores[1]],
        },
        {
            "nombre": "NVR 8 Canales",
            "categoria": categorias["Grabadores y NVR"],
            "fabricante": fabricantes["Hikvision"],
            "cantidad": 15,
            "bodega": bodega_principal,
            "proveedores": [proveedores[0]],
        },
        {
            "nombre": "NVR 16 Canales",
            "categoria": categorias["Grabadores y NVR"],
            "fabricante": fabricantes["Hikvision"],
            "cantidad": 10,
            "bodega": bodega_principal,
            "proveedores": [proveedores[0]],
        },
        {
            "nombre": "DVR 16 Canales",
            "categoria": categorias["Grabadores y NVR"],
            "fabricante": fabricantes["Dahua"],
            "cantidad": 12,
            "bodega": bodega_principal,
            "proveedores": [proveedores[1]],
        },
        {
            "nombre": "Switch PoE 8 Puertos",
            "categoria": categorias["Redes y Conectividad"],
            "fabricante": fabricantes["TP-Link"],
            "cantidad": 20,
            "bodega": bodega_principal,
            "proveedores": [proveedores[5], proveedores[8]],
        },
        {
            "nombre": "Switch PoE 24 Puertos",
            "categoria": categorias["Redes y Conectividad"],
            "fabricante": fabricantes["Cisco"],
            "cantidad": 6,
            "bodega": bodega_principal,
            "proveedores": [proveedores[5]],
        },
        {
            "nombre": "Router Gigabit",
            "categoria": categorias["Redes y Conectividad"],
            "fabricante": fabricantes["TP-Link"],
            "cantidad": 22,
            "bodega": bodega_principal,
            "proveedores": [proveedores[5]],
        },
        {
            "nombre": "Access Point AC",
            "categoria": categorias["Redes y Conectividad"],
            "fabricante": fabricantes["Ubiquiti"],
            "cantidad": 18,
            "bodega": bodega_principal,
            "proveedores": [proveedores[5], proveedores[8]],
        },
        {
            "nombre": "Controlador de Acceso 2 Puertas",
            "categoria": categorias["Control de Acceso"],
            "fabricante": fabricantes["Intelbras"],
            "cantidad": 8,
            "bodega": bodega_principal,
            "proveedores": [proveedores[6]],
        },
        {
            "nombre": "Lector RFID",
            "categoria": categorias["Control de Acceso"],
            "fabricante": fabricantes["Intelbras"],
            "cantidad": 40,
            "bodega": bodega_principal,
            "proveedores": [proveedores[6]],
        },
        {
            "nombre": "Boton de Salida",
            "categoria": categorias["Control de Acceso"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 50,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[6], proveedores[9]],
        },
        {
            "nombre": "Cerradura Electromagnetica 600lb",
            "categoria": categorias["Control de Acceso"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 20,
            "bodega": bodega_principal,
            "proveedores": [proveedores[6], proveedores[9]],
        },
        {
            "nombre": "Sirena 30W Exterior",
            "categoria": categorias["Alarmas e Intrusion"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 25,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[3], proveedores[9]],
        },
        {
            "nombre": "Panel de Alarma 8 Zonas",
            "categoria": categorias["Alarmas e Intrusion"],
            "fabricante": fabricantes["Intelbras"],
            "cantidad": 10,
            "bodega": bodega_principal,
            "proveedores": [proveedores[3], proveedores[6]],
        },
        {
            "nombre": "Sensor Movimiento PIR",
            "categoria": categorias["Alarmas e Intrusion"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 60,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[3], proveedores[9]],
        },
        {
            "nombre": "Sensor Magnetico",
            "categoria": categorias["Alarmas e Intrusion"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 70,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[3], proveedores[9]],
        },
        {
            "nombre": "UPS 1000VA",
            "categoria": categorias["Energia y UPS"],
            "fabricante": fabricantes["APC"],
            "cantidad": 8,
            "bodega": bodega_principal,
            "proveedores": [proveedores[7]],
        },
        {
            "nombre": "UPS 1500VA",
            "categoria": categorias["Energia y UPS"],
            "fabricante": fabricantes["APC"],
            "cantidad": 6,
            "bodega": bodega_principal,
            "proveedores": [proveedores[7]],
        },
        {
            "nombre": "Fuente 12V 5A",
            "categoria": categorias["Energia y UPS"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 30,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[7], proveedores[9]],
        },
        {
            "nombre": "Gabinete Rack 12U",
            "categoria": categorias["Accesorios y Montaje"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 8,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[9]],
        },
        {
            "nombre": "Gabinete Rack 6U",
            "categoria": categorias["Accesorios y Montaje"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 10,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[9]],
        },
        {
            "nombre": "Patch Panel 24 Puertos",
            "categoria": categorias["Accesorios y Montaje"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 20,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[9]],
        },
        {
            "nombre": "Cable UTP Cat6 305m",
            "categoria": categorias["Cableado y Conectores"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 50,
            "bodega": bodega_principal,
            "proveedores": [proveedores[5], proveedores[9]],
        },
        {
            "nombre": "Conector RJ45 (Bolsa 100u)",
            "categoria": categorias["Cableado y Conectores"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 80,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[9]],
        },
        {
            "nombre": "Canaleta 40x20",
            "categoria": categorias["Accesorios y Montaje"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 40,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[9]],
        },
        {
            "nombre": "Soporte Camara Universal",
            "categoria": categorias["Accesorios y Montaje"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 35,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[9]],
        },
        {
            "nombre": "Cable Coaxial 305m",
            "categoria": categorias["Cableado y Conectores"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 28,
            "bodega": bodega_principal,
            "proveedores": [proveedores[9]],
        },
        {
            "nombre": "Conector BNC (Pack 50u)",
            "categoria": categorias["Cableado y Conectores"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 45,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[9]],
        },
        {
            "nombre": "Video Balun Pasivo",
            "categoria": categorias["Cableado y Conectores"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 40,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[9]],
        },
        {
            "nombre": "Disco Duro 2TB Vigilancia",
            "categoria": categorias["Grabadores y NVR"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 18,
            "bodega": bodega_principal,
            "proveedores": [proveedores[2]],
        },
        {
            "nombre": "Disco Duro 4TB Vigilancia",
            "categoria": categorias["Grabadores y NVR"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 12,
            "bodega": bodega_principal,
            "proveedores": [proveedores[2]],
        },
        {
            "nombre": "Monitor 24 pulgadas",
            "categoria": categorias["Accesorios y Montaje"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 10,
            "bodega": bodega_principal,
            "proveedores": [proveedores[2]],
        },
        {
            "nombre": "Kit Videoportero IP",
            "categoria": categorias["Control de Acceso"],
            "fabricante": fabricantes["Hikvision"],
            "cantidad": 6,
            "bodega": bodega_principal,
            "proveedores": [proveedores[0], proveedores[6]],
        },
        {
            "nombre": "Boton de Panico",
            "categoria": categorias["Alarmas e Intrusion"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 25,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[3], proveedores[9]],
        },
        {
            "nombre": "Licencia VMS 16 Canales",
            "categoria": categorias["Grabadores y NVR"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 20,
            "bodega": bodega_principal,
            "proveedores": [proveedores[2]],
        },
        {
            "nombre": "Convertidor Media Fiber",
            "categoria": categorias["Redes y Conectividad"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 14,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[8], proveedores[9]],
        },
        {
            "nombre": "Fibra Optica 1km",
            "categoria": categorias["Redes y Conectividad"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 8,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[8]],
        },
        {
            "nombre": "Fuente 12V 2A",
            "categoria": categorias["Energia y UPS"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 35,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[7], proveedores[9]],
        },
        {
            "nombre": "Cable Patch Cord Cat6 2m",
            "categoria": categorias["Cableado y Conectores"],
            "fabricante": fabricantes["Genérico"],
            "cantidad": 60,
            "bodega": bodega_secundaria,
            "proveedores": [proveedores[9]],
        },
    ]

    items_creados = []
    for data in items_data:
        item, _ = ItemEmpresa.objects.get_or_create(
            empresa=empresa,
            nombre=data["nombre"],
            defaults={
                "categoria": data["categoria"],
                "fabricante": data["fabricante"],
                "descripcion_corta": data["nombre"],
            },
        )
        item.proveedores_empresa.set(data["proveedores"])
        items_creados.append(item)
        stock, created = StockItemEnBodega.objects.get_or_create(
            item=item,
            defaults={"bodega": data["bodega"], "cantidad": data["cantidad"], "pmp": 0},
        )
        updated = False
        if not created and stock.bodega_id != data["bodega"].id:
            stock.bodega = data["bodega"]
            updated = True
        if stock.cantidad < data["cantidad"]:
            stock.cantidad = data["cantidad"]
            updated = True
        if updated:
            stock.save(update_fields=["bodega", "cantidad"])

    return items_creados


def ensure_software_catalog() -> List[object]:
    from core.models import Software

    software_names = [
        "Microsoft Office",
        "Google Chrome",
        "Mozilla Firefox",
        "Adobe Acrobat Reader",
        "WinRAR",
        "7-Zip",
        "VLC Media Player",
        "Zoom",
        "Microsoft Teams",
        "Slack",
        "AutoCAD",
        "Photoshop",
        "Visual Studio Code",
        "Python",
        "Node.js",
    ]
    softwares = []
    for nombre in software_names:
        software, _ = Software.objects.get_or_create(nombre=nombre)
        softwares.append(software)
    return softwares


def ensure_software_de_empresa(softwares: List[object], empresas: List[object]) -> None:
    from recursos.models import SoftwareDeEmpresa

    for empresa in empresas:
        for software in softwares:
            SoftwareDeEmpresa.objects.get_or_create(
                software=software,
                empresa=empresa,
                defaults={"activo": True},
            )


def ensure_equipos(
    clientes: List[Dict[str, object]],
    tecnico_usuario: object,
) -> List[object]:
    from recursos.models import AlmacenamientoEquipo, Equipo, UsuarioEquipo

    equipos_creados: List[object] = []
    equipos_data = [
        {
            "nombre_equipo": "Portatil HP ProBook 440",
            "tipo_equipo": "PORTATIL",
            "marca": "HP",
            "modelo": "ProBook 440",
            "ram": "16GB",
            "sistema_operativo": "WINDOWS10",
            "almacenamientos": ["SSD_256GB"],
        },
        {
            "nombre_equipo": "Escritorio Dell OptiPlex 7090",
            "tipo_equipo": "ESCRITORIO",
            "marca": "DELL",
            "modelo": "OptiPlex 7090",
            "ram": "32GB",
            "sistema_operativo": "WINDOWS11",
            "almacenamientos": ["SSD_512GB"],
        },
        {
            "nombre_equipo": "Portatil Lenovo ThinkPad T14",
            "tipo_equipo": "PORTATIL",
            "marca": "LENOVO",
            "modelo": "ThinkPad T14",
            "ram": "16GB",
            "sistema_operativo": "WINDOWS11",
            "almacenamientos": ["SSD_256GB", "HDD_1TB"],
        },
    ]

    for cliente in clientes:
        empresa = cliente["empresa"]
        usuarios = cliente["usuarios"]
        for idx, data in enumerate(equipos_data, start=1):
            numero_serie = f"{empresa.rut_empresa}-{idx:03d}"
            equipo, _ = Equipo.objects.get_or_create(
                numero_serie=numero_serie,
                defaults={
                    "nombre_equipo": data["nombre_equipo"],
                    "cliente": empresa,
                    "registrado_por": tecnico_usuario,
                    "tipo_equipo": data["tipo_equipo"],
                    "marca": data["marca"],
                    "modelo": data["modelo"],
                    "ram": data["ram"],
                    "sistema_operativo": data["sistema_operativo"],
                },
            )
            equipos_creados.append(equipo)

            for almacenamiento in data["almacenamientos"]:
                AlmacenamientoEquipo.objects.get_or_create(
                    equipo=equipo,
                    almacenamiento=almacenamiento,
                    defaults={
                        "adicional": almacenamiento != data["almacenamientos"][0]
                    },
                )

            if usuarios:
                if idx == 1:
                    UsuarioEquipo.objects.get_or_create(
                        equipo=equipo,
                        usuario=usuarios[0],
                        defaults={"estado": True},
                    )
                elif idx == 2 and len(usuarios) > 1:
                    UsuarioEquipo.objects.get_or_create(
                        equipo=equipo,
                        usuario=usuarios[1],
                        defaults={
                            "estado": False,
                            "fecha_devolucion": date.today() - timedelta(days=30),
                        },
                    )
    return equipos_creados


def ensure_software_instalado(equipos: List[object], softwares: List[object]) -> None:
    from core.models import Software
    from django.contrib.contenttypes.models import ContentType
    from recursos.models import SoftwareInstalado

    if not equipos or not softwares:
        return

    content_type = ContentType.objects.get_for_model(Software)
    for equipo in equipos:
        for software in softwares[:3]:
            SoftwareInstalado.objects.get_or_create(
                equipo=equipo,
                content_type=content_type,
                software_id=software.id,
                defaults={"version": "latest"},
            )


def ensure_servicios_catalogo() -> None:
    from contratos.models import CaracteristicaServicio, PlanServicio, Servicio

    caracteristicas = []
    for nombre in [
        "Incluye materiales",
        "24/7 Disponibilidad",
        "Garantia extendida",
        "Respuesta prioritaria",
        "Informe tecnico",
    ]:
        caracteristica, _ = CaracteristicaServicio.objects.get_or_create(nombre=nombre)
        caracteristicas.append(caracteristica)

    servicios_data = [
        ("Mantencion Preventiva de Infraestructura", "mantencion"),
        ("Desarrollo de Aplicacion Web Personalizada", "desarrollo"),
        ("Soporte Tecnico Nivel 2", "soporte"),
        ("Capacitacion en Nuevas Tecnologias", "capacitacion"),
        ("Hosting y Almacenamiento en Datacenter", "datacenter"),
        ("Migracion de Sistemas Legacy", "desarrollo"),
        ("Monitoreo 24/7 de Infraestructura", "datacenter"),
    ]

    servicios = []
    for nombre, categoria in servicios_data:
        servicio, _ = Servicio.objects.get_or_create(
            nombre=nombre,
            defaults={"categoria": categoria, "descripcion": nombre},
        )
        servicio.caracteristicas.set(caracteristicas)
        servicios.append(servicio)

    planes_data = [
        ("Plan Basico de Mantenimiento", [servicios[0]]),
        ("Plan Completo de Soporte", [servicios[2]]),
        (
            "Plan Empresarial Premium",
            [servicios[0], servicios[2], servicios[4]],
        ),
    ]

    for nombre, servicios_plan in planes_data:
        plan, _ = PlanServicio.objects.get_or_create(nombre=nombre)
        plan.servicios.set(servicios_plan)


def ensure_contratos_catalogo() -> None:
    from contratos.models import CondicionEspecial, Licencia, Visita

    visitas = [
        "Visita de Mantenimiento Mensual",
        "Visita de Mantenimiento Trimestral",
        "Visita de Mantenimiento Semestral",
        "Visita de Mantenimiento Anual",
        "Visita de Soporte Tecnico",
        "Visita de Inspeccion de Equipos",
        "Visita de Instalacion de Software",
        "Visita de Capacitacion de Usuarios",
    ]
    for descripcion in visitas:
        Visita.objects.get_or_create(descripcion=descripcion)

    licencias = [
        ("Microsoft 365 Business Standard", "Microsoft"),
        ("Microsoft 365 E3", "Microsoft"),
        ("AutoCAD", "Autodesk"),
        ("Adobe Creative Cloud", "Adobe"),
        ("Slack Business+", "Slack"),
        ("Zoom Business", "Zoom"),
        ("Antivirus Corporativo", "Kaspersky/ESET/Symantec"),
    ]
    for nombre, proveedor in licencias:
        Licencia.objects.get_or_create(nombre=nombre, proveedor=proveedor)

    condiciones = [
        "SLA 4 horas",
        "Soporte 24/7",
        "Garantia extendida 3 anos",
        "Capacitacion incluida",
        "Respaldo de datos diario",
        "Mantenimiento preventivo trimestral",
        "Reemplazo de equipos en caso de falla",
        "Acceso prioritario a nuevas funcionalidades",
    ]
    for titulo in condiciones:
        CondicionEspecial.objects.get_or_create(titulo=titulo, descripcion=titulo)


def ensure_rendiciones_categorias() -> None:
    from rendiciones.models import CategoriaGastoRendicion

    categorias = [
        "Combustible",
        "Peaje",
        "Estacionamiento",
        "Taxi/Uber",
        "Transporte Publico (Metro/Bus)",
        "Arriendo de Vehiculo",
        "Desayuno / Almuerzo / Cena",
        "Colacion",
        "Hotel",
        "Hostal",
        "Cables y Conectores",
        "Herramientas",
        "Material Electrico",
        "Tornilleria",
        "Consumibles",
        "Llamadas Telefonicas",
        "Internet Movil",
        "Capacitacion",
        "Impresiones y Fotocopias",
        "Envio de Documentos",
        "Gastos Varios",
    ]
    for nombre in categorias:
        CategoriaGastoRendicion.objects.get_or_create(nombre=nombre)


def ensure_acuerdos_confidencialidad() -> None:
    from core.models import AcuerdoConfidencialidadBase

    AcuerdoConfidencialidadBase.objects.get_or_create(
        titulo="NDA Estandar",
        defaults={
            "contenido": "Este acuerdo protege la confidencialidad entre las partes.",
        },
    )


# FUNCIÓN ELIMINADA: ensure_cotizaciones
# Las cotizaciones deben crearse manualmente por el usuario para pruebas reales
            "tipo_moneda": "2",
            "porcentaje_recargo": 24,
        },
        {
            "cliente_idx": 9,
            "nombre": "Cotizacion Licencias VMS",
            "descripcion": "Licenciamiento y monitoreo",
            "estado": "enviada",
            "items_indices": [39, 36],
            "cantidades": [10, 1],
            "dias_vencimiento": 30,
            "tipo_moneda": "2",
            "porcentaje_recargo": 19,
        },
    ]

    cotizaciones_creadas = []
    for idx, cot_data in enumerate(cotizaciones_data):
        if cot_data["cliente_idx"] >= len(clientes):
            continue

        cliente = clientes[cot_data["cliente_idx"]]["empresa"]
        fecha_venc = date.today() + timedelta(days=cot_data["dias_vencimiento"])

        cotizacion, created = Cotizacion.objects.get_or_create(
            empresa=empresa,
            cliente=cliente,
            nombre=cot_data["nombre"],
            defaults={
                "descripcion": cot_data["descripcion"],
                "estado": cot_data["estado"],
                "fecha_vencimiento": fecha_venc,
                "tipo_moneda": cot_data["tipo_moneda"],
                "porcentaje_recargo": cot_data["porcentaje_recargo"],
                "dolar_observado": Decimal("950.00"),
                "valor_uf": Decimal("37500.00"),
                "fecha_tipo_cambio": date.today(),
            },
        )

        if created:
            # Agregar items a la cotización
            for item_idx, cantidad in zip(
                cot_data["items_indices"], cot_data["cantidades"]
            ):
                if item_idx < len(items):
                    item = items[item_idx]
                    # Precio base simulado
                    precio_base = Decimal("50000") * (item_idx + 1)

                    ItemCotizacion.objects.create(
                        cotizacion=cotizacion,
                        item_empresa=item,
                        nombre=item.nombre,
                        descripcion=item.descripcion_corta or item.nombre,
                        cantidad=cantidad,
                        precio_unitario=precio_base,
                    )

        cotizaciones_creadas.append(cotizacion)

    return cotizaciones_creadas


# FUNCIÓN ELIMINADA: ensure_ordenes_trabajo
# Las órdenes de trabajo deben crearse manualmente por el usuario para pruebas reales


# FUNCIÓN ELIMINADA: ensure_contratos
# Los contratos deben crearse manualmente por el usuario para pruebas reales


# FUNCIÓN ELIMINADA: ensure_ordenes_compra
# Las órdenes de compra deben crearse manualmente por el usuario para pruebas reales


# FUNCIÓN ELIMINADA: ensure_compras_guias_movimientos
# Las compras, guías de salida y movimientos de stock deben crearse manualmente
# por el usuario para pruebas reales de flujos operativos


def run_seed() -> None:
    bootstrap_django()
    ensure_migrations()

    from django.db import transaction

    with transaction.atomic():
        grupos = ensure_groups()
        empresa_base, sucursal_base = ensure_empresa_base()
        internos = ensure_internal_users(empresa_base, sucursal_base, grupos)
        clientes = ensure_client_companies(empresa_base, grupos)

        bodegas = ensure_bodegas(empresa_base, sucursal_base)
        categorias, fabricantes, proveedores = ensure_catalogs(empresa_base)
        items = ensure_items_and_stock(
            empresa_base, bodegas, categorias, fabricantes, proveedores
        )

        softwares = ensure_software_catalog()
        empresas_para_software = [empresa_base] + [c["empresa"] for c in clientes]
        ensure_software_de_empresa(softwares, empresas_para_software)

        equipos = ensure_equipos(clientes, internos["tecnico"])
        ensure_software_instalado(equipos, softwares)

        ensure_servicios_catalogo()
        ensure_contratos_catalogo()
        ensure_rendiciones_categorias()
        ensure_acuerdos_confidencialidad()

    print("\n" + "=" * 80)
    print("Seed base completado exitosamente!".center(80))
    print("=" * 80)
    print("\n📦 Datos base creados:")
    print("   ✅ Empresa base (Snabbit) con recargo y PPM variables")
    print("   ✅ Sucursal base con datos de ubicación completos")
    print("   ✅ Grupos de permisos")
    print("   ✅ Usuarios internos (Admin, Ventas, Bodega, Contabilidad, Supervisor, Técnico)")
    print(f"   ✅ {len(clientes)} Empresas cliente con recargo y PPM variables")
    print("   ✅ Relaciones empresa-cliente")
    print("   ✅ Categorías, Fabricantes y Proveedores")
    print(f"   ✅ {len(items)} Items con stock en bodegas")
    print("   ✅ Catálogo de software y equipos")
    print("   ✅ Catálogos de servicios, contratos y rendiciones")
    print("\n📝 Nota: Cotizaciones, OT, Contratos, OC, Compras y Guías")
    print("   NO se crean automáticamente para permitir pruebas manuales reales.")
    print()


def main() -> int:
    print_header("SEED BASE - Poblando datos base del sistema")

    print("\n> Scripts a ejecutar:")
    for i, script in enumerate(SEED_SCRIPTS, 1):
        status = "Requerido" if script["required"] else "Opcional"
        condition_text = (
            f" (Condicion: {script['condition']})" if script["condition"] else ""
        )
        print(
            f"   {i}. {script['name']:30} - {script['description']} [{status}]{condition_text}"
        )

    missing_scripts = []
    for script in SEED_SCRIPTS:
        script_path = SCRIPTS_DIR / script["name"]
        if not script_path.exists():
            missing_scripts.append(script["name"])

    if missing_scripts:
        print("\nERROR: Los siguientes scripts no existen:")
        for script_name in missing_scripts:
            print(f"   - {script_name}")
        print(f"\nRuta esperada: {SCRIPTS_DIR}")
        return 1

    total_scripts = len(SEED_SCRIPTS)
    successful = 0
    failed = []
    skipped = []

    for i, script in enumerate(SEED_SCRIPTS, 1):
        script_path = SCRIPTS_DIR / script["name"]

        if script["condition"] and not check_condition(script["condition"]):
            print(
                f"\nSaltando {script['name']} - Condicion '{script['condition']}' no cumplida"
            )
            skipped.append(script["name"])
            continue

        print_step(i, total_scripts, script["name"], script["description"])

        if run_script(script_path):
            print(f"\nOK: {script['name']} completado exitosamente")
            successful += 1
        else:
            print(f"\nERROR en {script['name']}")
            failed.append(script["name"])
            if script["required"]:
                print("Este script es REQUERIDO. Deteniendo ejecucion.")
                break

    print_header("RESUMEN DE EJECUCION")
    print("\nEstadisticas:")
    print(f"   - Scripts ejecutados:    {successful}/{total_scripts}")
    print(f"   - Scripts exitosos:      {successful}")
    print(f"   - Scripts fallidos:      {len(failed)}")
    print(f"   - Scripts saltados:      {len(skipped)}")

    if failed:
        print("\nScripts que fallaron:")
        for script_name in failed:
            print(f"   - {script_name}")
        return 1

    print("\nSEED BASE FINALIZADO CON EXITO")
    return 0


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--internal":
        run_seed()
        sys.exit(0)
    sys.exit(main())
