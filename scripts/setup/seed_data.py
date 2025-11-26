#!/usr/bin/env python
"""\
Script para poblar la base de datos con datos de prueba.

Qué hace:
- Crea usuarios internos con diferentes roles
- Carga empresas y usuarios cliente desde Excel o usa datos de prueba como respaldo
- Registra equipos para clientes y realiza asignaciones parciales
- Crea items de prueba (productos/servicios)
- Crea bodegas con stock
- Crea cotizaciones de ejemplo
- Útil para desarrollo y testing

Cuándo usar:
- Después de reset_db para tener datos de prueba
- Testing de funcionalidades con datos realistas
- Demos del sistema

Prerequisitos:
- Base de datos migrada
- Superusuario configurado (setup_superuser.py)

Uso:
    cd backend
    backend\\ENV\\Scripts\\python.exe ..\\scripts\\setup\\seed_data.py
"""
import os
import sys
from datetime import date, timedelta
from pathlib import Path

import django

# Setup Django
backend_path = os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)
sys.path.insert(0, os.path.join(backend_path, "backend"))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "sw_erp.settings")
django.setup()

from bodegas.models import Bodega
from core.models import PersonalizacionUsuario
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from empresas.models import Empresa, SucursalEmpresa, UsuarioEmpresa
from items.models import Categoria, Fabricante, ItemEmpresa
from openpyxl import load_workbook
from recursos.models import Equipo, UsuarioEquipo

User = get_user_model()

BASE_BACKEND_PATH = Path(backend_path) / "backend"
EXCEL_USER_FILES = [
    "usuarios_aygasociados.xlsx",
    "usuarios_camacoes.xlsx",
    "usuarios_molinarios.xlsx",
    "usuarios_prodalmen.xlsx",
]
DEFAULT_EXCEL_PASSWORD = "test1234"



def crear_usuarios_prueba():
    """Crea usuarios de prueba con diferentes roles."""
    usuarios_data = [
        {
            "email": "tecnico@snabbit.cl",
            "first_name": "Juan",
            "last_name": "Técnico",
            "password": "test1234",
            "grupos": ["tecnico"],
        },
        {
            "email": "bodeguero@snabbit.cl",
            "first_name": "María",
            "last_name": "Bodeguera",
            "password": "test1234",
            "grupos": ["bodeguero"],
        },
        {
            "email": "admin@snabbit.cl",
            "first_name": "Pedro",
            "last_name": "Admin",
            "password": "test1234",
            "grupos": ["staff", "superadmin"],
        },
    ]

    empresa_snabbit = Empresa.objects.get(rut_empresa="11111111-1")
    sucursal = SucursalEmpresa.objects.filter(empresa=empresa_snabbit).first()

    usuarios = []
    for data in usuarios_data:
        user, created = User.objects.get_or_create(
            email=data["email"],
            defaults={
                "first_name": data["first_name"],
                "last_name": data["last_name"],
            },
        )

        if created:
            user.set_password(data["password"])
            user.save()
            print(f"✓ Usuario '{user.email}' creado")
        else:
            print(f"  Usuario '{user.email}' ya existe")

        # Asociar a empresa con grupos
        usuario_empresa, _ = UsuarioEmpresa.objects.get_or_create(
            usuario=user,
            defaults={
                "sucursal": sucursal,
                "estado": "1",
            },
        )
        if usuario_empresa.sucursal_id != sucursal.id:
            usuario_empresa.sucursal = sucursal
            usuario_empresa.save(update_fields=["sucursal"])

        # Asignar grupos
        grupos = Group.objects.filter(name__in=data["grupos"])
        usuario_empresa.grupos.set(grupos)

        usuarios.append(user)

    return usuarios


def crear_categorias_y_fabricantes():
    """Crea categorías y fabricantes de ejemplo."""
    categorias_data = [
        "Cámaras de Seguridad",
        "DVR/NVR",
        "Alarmas",
        "Control de Acceso",
        "Cables y Conectores",
    ]

    categorias = []
    for nombre in categorias_data:
        categoria, created = Categoria.objects.get_or_create(nombre=nombre)
        categorias.append(categoria)
        if created:
            print(f"✓ Categoría '{categoria.nombre}' creada")

    fabricantes_data = [
        "Hikvision",
        "Dahua",
        "Samsung",
        "Axis",
        "Genérico",
    ]

    fabricantes = []
    for nombre in fabricantes_data:
        fabricante, created = Fabricante.objects.get_or_create(nombre=nombre)
        fabricantes.append(fabricante)
        if created:
            print(f"✓ Fabricante '{fabricante.nombre}' creado")

    return categorias, fabricantes


def crear_items_prueba(categorias, fabricantes):
    """Crea items de prueba."""
    empresa_snabbit = Empresa.objects.get(rut_empresa="11111111-1")

    items_data = [
        {
            "nombre": "Cámara Domo 2MP",
            "descripcion_corta": "CAM-DOMO-001",
            "categoria": categorias[0],
            "fabricante": fabricantes[0],
        },
        {
            "nombre": "DVR 8 Canales",
            "descripcion_corta": "DVR-8CH-001",
            "categoria": categorias[1],
            "fabricante": fabricantes[1],
        },
        {
            "nombre": "Cable UTP Cat5e",
            "descripcion_corta": "CAB-UTP-001",
            "categoria": categorias[4],
            "fabricante": fabricantes[4],
        },
    ]

    items = []
    for data in items_data:
        item, created = ItemEmpresa.objects.get_or_create(
            empresa=empresa_snabbit,
            nombre=data["nombre"],
            defaults={
                "descripcion_corta": data["descripcion_corta"],
                "categoria": data["categoria"],
                "fabricante": data["fabricante"],
            },
        )
        items.append(item)
        if created:
            print(f"✓ Item '{item.nombre}' creado")
        else:
            cambios = False
            if data["categoria"] and item.categoria_id != data["categoria"].id:
                item.categoria = data["categoria"]
                cambios = True
            if data["fabricante"] and item.fabricante_id != data["fabricante"].id:
                item.fabricante = data["fabricante"]
                cambios = True
            if (
                data["descripcion_corta"]
                and item.descripcion_corta != data["descripcion_corta"]
            ):
                item.descripcion_corta = data["descripcion_corta"]
                cambios = True
            if cambios:
                item.save()
                print(f"✓ Item '{item.nombre}' actualizado")

    return items


def crear_bodegas_prueba():
    """Crea bodegas de prueba."""
    empresa_snabbit = Empresa.objects.get(rut_empresa="11111111-1")
    sucursal = _obtener_sucursal_principal(empresa_snabbit)

    bodegas_data = [
        {
            "nombre": "Bodega Principal",
        },
        {
            "nombre": "Bodega Secundaria",
        },
    ]

    bodegas = []
    for data in bodegas_data:
        bodega, created = Bodega.objects.get_or_create(
            nombre=data["nombre"],
            sucursal=sucursal,
        )
        bodegas.append(bodega)
        if created:
            print(f"✓ Bodega '{bodega.nombre}' creada")
        elif bodega.sucursal_id != sucursal.id:
            bodega.sucursal = sucursal
            bodega.save(update_fields=["sucursal"])
            print(f"✓ Bodega '{bodega.nombre}' actualizada")

    return bodegas


def _slugify(nombre: str) -> str:
    return "".join(char for char in nombre.lower() if char.isalnum())


def _clean_str(value):
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _obtener_sucursal_principal(empresa: Empresa) -> SucursalEmpresa:
    sucursal = empresa.sucursales.first()
    if not sucursal:
        sucursal = SucursalEmpresa.objects.create(
            empresa=empresa,
            nombre="Casa Matriz",
            direccion=empresa.direccion_principal or "Dirección no especificada",
        )
    return sucursal


def preparar_empresas_y_usuarios_cliente():
    """Carga usuarios y empresas desde Excel, con fallback a datos de prueba."""

    print("--- Preparando empresas y usuarios cliente ---")

    archivos = []
    for nombre in EXCEL_USER_FILES:
        ruta = BASE_BACKEND_PATH / nombre
        if ruta.exists():
            archivos.append(ruta)
        else:
            print(f"  ⚠️ Archivo '{nombre}' no encontrado, se omite.")

    totales = {"empresas": 0, "usuarios": 0, "usando_excel": False}
    usuarios: list[UsuarioEmpresa] = []
    empresas_dict: dict[int, Empresa] = {}

    if archivos:
        totales["usando_excel"] = True
        grupo_representante = Group.objects.filter(name="representante_legal").first()

        for archivo in archivos:
            wb = load_workbook(archivo)
            sheet = wb.active
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

            try:
                idx_empresa = headers.index("Empresa")
                idx_nombre = headers.index("Nombre")
                idx_apellido = headers.index("Apellido")
                idx_correo = headers.index("Correo")
            except ValueError:
                print(f"  ⚠️ Encabezados inesperados en '{archivo.name}', se omite.")
                continue

            print(f"  Procesando '{archivo.name}' ({sheet.max_row - 1} filas)")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                empresa_nombre = _clean_str(row[idx_empresa])
                email = _clean_str(row[idx_correo]).lower()
                first_name = _clean_str(row[idx_nombre])
                last_name = _clean_str(row[idx_apellido])

                if not email:
                    continue

                if not empresa_nombre:
                    empresa_nombre = "Empresa sin nombre"

                empresa_defaults = {
                    "direccion_principal": "Dirección no especificada",
                    "telefono": "",
                    "email": "",
                }
                empresa, creada = Empresa.objects.get_or_create(
                    nombre=empresa_nombre,
                    defaults=empresa_defaults,
                )
                empresas_dict[empresa.id] = empresa
                if creada:
                    totales["empresas"] += 1

                sucursal = _obtener_sucursal_principal(empresa)

                user, creado_usuario = User.objects.get_or_create(
                    email=email,
                    defaults={
                        "first_name": first_name or "Usuario",
                        "last_name": last_name or "",
                        "is_active": True,
                    },
                )
                if creado_usuario:
                    user.set_password(DEFAULT_EXCEL_PASSWORD)
                    user.is_active = True
                    user.save()
                    totales["usuarios"] += 1
                else:
                    actualizado = False
                    if first_name and user.first_name != first_name:
                        user.first_name = first_name
                        actualizado = True
                    if last_name and user.last_name != last_name:
                        user.last_name = last_name
                        actualizado = True
                    if not user.is_active:
                        user.is_active = True
                        actualizado = True
                    if actualizado:
                        user.save()

                usuario_empresa, creado_ue = UsuarioEmpresa.objects.get_or_create(
                    usuario=user,
                    defaults={
                        "sucursal": sucursal,
                        "estado": "1",
                    },
                )

                if not creado_ue:
                    cambios = False
                    if usuario_empresa.sucursal_id != sucursal.id:
                        usuario_empresa.sucursal = sucursal
                        cambios = True
                    if usuario_empresa.estado != "1":
                        usuario_empresa.estado = "1"
                        cambios = True
                    if cambios:
                        usuario_empresa.save()
                else:
                    usuario_empresa.save()

                if grupo_representante:
                    usuario_empresa.grupos.add(grupo_representante)

                personalizacion, creada_personalizacion = (
                    PersonalizacionUsuario.objects.get_or_create(
                        usuario=user,
                        defaults={
                            "tema": "3",
                            "font_size": 14,
                            "sucursal_principal": sucursal,
                        },
                    )
                )
                if (
                    not creada_personalizacion
                    and personalizacion.sucursal_principal_id != sucursal.id
                ):
                    personalizacion.sucursal_principal = sucursal
                    personalizacion.save(update_fields=["sucursal_principal"])

                usuarios.append(usuario_empresa)

    if usuarios:
        return list(empresas_dict.values()), usuarios, totales

    totales["usando_excel"] = False
    print("  ⚠️ No se encontraron planillas válidas. Usando datos de prueba.")

    empresas_data = [
        {
            "rut": "76123456-7",
            "nombre": "Empresa Cliente A",
            "direccion_principal": "Av. Principal 100",
            "telefono": "+56922334455",
            "email": "contacto@clientea.cl",
        },
        {
            "rut": "76234567-8",
            "nombre": "Empresa Cliente B",
            "direccion_principal": "Calle Secundaria 200",
            "telefono": "+56933445566",
            "email": "info@clienteb.cl",
        },
    ]

    empresas: list[Empresa] = []
    for data in empresas_data:
        empresa, created = Empresa.objects.get_or_create(
            rut_empresa=data["rut"],
            defaults={
                "nombre": data["nombre"],
                "direccion_principal": data["direccion_principal"],
                "telefono": data["telefono"],
                "email": data["email"],
            },
        )
        empresas.append(empresa)
        if created:
            print(f"✓ Empresa '{empresa.nombre}' creada")
        else:
            print(f"  Empresa '{empresa.nombre}' ya existe")

        _obtener_sucursal_principal(empresa)

    usuarios_existentes: list[UsuarioEmpresa] = []
    for empresa in empresas:
        sucursal = _obtener_sucursal_principal(empresa)
        existentes = list(
            UsuarioEmpresa.objects.filter(
                sucursal__empresa=empresa, estado="1"
            ).select_related("usuario")
        )
        usuarios_existentes.extend(existentes)

        faltantes = max(0, 2 - len(existentes))
        if faltantes == 0:
            continue

        dominio = _slugify(empresa.nombre) or "cliente"
        base_datos = [
            {
                "email": f"soporte@{dominio}.cl",
                "first_name": "Soporte",
                "last_name": empresa.nombre,
            },
            {
                "email": f"backoffice@{dominio}.cl",
                "first_name": "Backoffice",
                "last_name": empresa.nombre,
            },
        ]

        for data in base_datos:
            if faltantes <= 0:
                break

            user, created = User.objects.get_or_create(
                email=data["email"],
                defaults={
                    "first_name": data["first_name"],
                    "last_name": data["last_name"],
                    "is_active": True,
                },
            )
            if created:
                user.set_password(DEFAULT_EXCEL_PASSWORD)
                user.save()
                print(f"✓ Usuario cliente creado: {user.email}")
            else:
                print(f"  Usuario cliente reutilizado: {user.email}")

            usuario_empresa, _ = UsuarioEmpresa.objects.get_or_create(
                usuario=user,
                defaults={"sucursal": sucursal, "estado": "1"},
            )
            if usuario_empresa.sucursal_id != sucursal.id:
                usuario_empresa.sucursal = sucursal
                usuario_empresa.save(update_fields=["sucursal"])

            usuarios_existentes.append(usuario_empresa)
            faltantes -= 1

    return empresas, usuarios_existentes, totales


def crear_equipos_para_empresas(
    empresas: list[Empresa], registrado_por: UsuarioEmpresa | None
) -> list[Equipo]:
    if not registrado_por:
        print("⚠️ No se encontró el usuario técnico para registrar equipos.")
        return []

    equipos_creados: list[Equipo] = []
    base_equipos = [
        {
            "tipo_equipo": "PORTATIL",
            "marca": "HP",
            "modelo": "ProBook 440",
            "ram": "16",
            "sistema_operativo": "WINDOWS",
        },
        {
            "tipo_equipo": "ESCRITORIO",
            "marca": "Dell",
            "modelo": "OptiPlex 7090",
            "ram": "32",
            "sistema_operativo": "WINDOWS",
        },
        {
            "tipo_equipo": "PORTATIL",
            "marca": "Lenovo",
            "modelo": "ThinkPad T14",
            "ram": "16",
            "sistema_operativo": "WINDOWS",
        },
    ]

    for empresa in empresas:
        dominio = _slugify(empresa.nombre).upper() or "CLI"
        for idx, datos_equipo in enumerate(base_equipos, start=1):
            numero_serie = f"{dominio}-{idx:03d}"
            equipo, created = Equipo.objects.get_or_create(
                numero_serie=numero_serie,
                defaults={
                    "cliente": empresa,
                    "registrado_por": registrado_por,
                    **datos_equipo,
                },
            )
            equipos_creados.append(equipo)
            if created:
                print(f"✓ Equipo creado: {equipo.numero_serie} para {empresa.nombre}")
            else:
                print(f"  Equipo ya existe: {equipo.numero_serie} para {empresa.nombre}")

    return equipos_creados


def crear_asignaciones_equipos(
    equipos: list[Equipo], usuarios_cliente: list[UsuarioEmpresa]
) -> list[UsuarioEquipo]:
    asignaciones: list[UsuarioEquipo] = []
    usuarios_por_empresa: dict[int, list[UsuarioEmpresa]] = {}
    for usuario in usuarios_cliente:
        empresa_id = usuario.sucursal.empresa_id
        usuarios_por_empresa.setdefault(empresa_id, []).append(usuario)

    equipos_por_empresa: dict[int, list[Equipo]] = {}
    for equipo in equipos:
        equipos_por_empresa.setdefault(equipo.cliente_id, []).append(equipo)

    for empresa_id, equipos_empresa in equipos_por_empresa.items():
        usuarios = usuarios_por_empresa.get(empresa_id, [])
        if not usuarios:
            print("⚠️ No hay usuarios cliente para asignar equipos en la empresa", empresa_id)
            continue

        activos: list[tuple[Equipo, UsuarioEmpresa]] = []
        devueltos: list[tuple[Equipo, UsuarioEmpresa]] = []

        if usuarios:
            activos.append((equipos_empresa[0], usuarios[0]))
        if len(usuarios) > 1 and len(equipos_empresa) > 1:
            devueltos.append((equipos_empresa[1], usuarios[1]))

        for equipo, usuario in activos:
            usuario_equipo, _ = UsuarioEquipo.objects.get_or_create(
                equipo=equipo,
                usuario=usuario,
                defaults={"observaciones": "Equipo asignado para soporte"},
            )
            asignaciones.append(usuario_equipo)
            print(
                f"✓ Equipo activo asignado: {equipo.numero_serie} -> {usuario.usuario.email}"
            )

        for equipo, usuario in devueltos:
            usuario_equipo, _ = UsuarioEquipo.objects.get_or_create(
                equipo=equipo,
                usuario=usuario,
                defaults={
                    "estado": False,
                    "fecha_devolucion": date.today() - timedelta(days=15),
                    "observaciones": "Equipo devuelto para mantenimiento",
                },
            )
            asignaciones.append(usuario_equipo)
            print(
                f"✓ Equipo devuelto registrado: {equipo.numero_serie} -> {usuario.usuario.email}"
            )

    return asignaciones


def main():
    print("=" * 60)
    print("Población de Datos de Prueba")
    print("=" * 60)
    print()

    # Verificar que existe empresa base
    try:
        Empresa.objects.get(rut_empresa="11111111-1")
    except Empresa.DoesNotExist:
        print("❌ Error: No se encontró la empresa base 'Snabbit'.")
        print("   Ejecuta primero: setup_superuser.py")
        return

    print("--- Creando usuarios de prueba ---")
    usuarios = crear_usuarios_prueba()
    print()

    empresas, usuarios_clientes, totales_clientes = (
        preparar_empresas_y_usuarios_cliente()
    )
    print()

    registrado_por = UsuarioEmpresa.objects.filter(
        usuario__email="tecnico@snabbit.cl"
    ).first()
    print("--- Registrando equipos de clientes ---")
    equipos = crear_equipos_para_empresas(empresas, registrado_por)
    print()

    print("--- Asignando equipos a usuarios cliente ---")
    asignaciones = crear_asignaciones_equipos(equipos, usuarios_clientes)
    print()

    print("--- Creando categorías y fabricantes ---")
    categorias, fabricantes = crear_categorias_y_fabricantes()
    print()

    print("--- Creando items de prueba ---")
    items = crear_items_prueba(categorias, fabricantes)
    print()

    print("--- Creando bodegas de prueba ---")
    bodegas = crear_bodegas_prueba()
    print()

    print("=" * 60)
    print("✓ Datos de prueba creados exitosamente")
    print("=" * 60)
    print()
    print("Resumen:")
    print(f"- Empresas: {len(empresas) + 1} (incluyendo Snabbit)")
    print(f"- Usuarios: {len(usuarios)}")
    print(f"- Usuarios cliente: {len(usuarios_clientes)}")
    print(f"- Categorías: {len(categorias)}")
    print(f"- Fabricantes: {len(fabricantes)}")
    print(f"- Items: {len(items)}")
    print(f"- Bodegas: {len(bodegas)}")
    print(f"- Equipos creados: {len(equipos)}")
    print(f"- Asignaciones de equipos: {len(asignaciones)}")
    print(f"- Empresas desde Excel: {totales_clientes['empresas']}")
    print(f"- Usuarios desde Excel: {totales_clientes['usuarios']}")
    if not totales_clientes["usando_excel"]:
        print("  (Fallback a datos de prueba internos)")
    print()
    print("Usuarios de prueba creados:")
    print("  - tecnico@snabbit.cl / test1234")
    print("  - bodeguero@snabbit.cl / test1234")
    print("  - admin@snabbit.cl / test1234")
    print()


if __name__ == "__main__":
    main()
